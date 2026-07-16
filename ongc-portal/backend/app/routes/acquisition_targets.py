from fastapi import APIRouter, Depends, HTTPException, Query, Body
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload
from app.database import get_db
from app.models.base import AcquisitionTarget, ManpowerEmployee, User, TargetMonthHistory
from app.auth.deps import get_current_user
from app.auth.security import verify_password
from typing import Optional
import io, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

router = APIRouter()

MONTH_COLS = ["apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec", "jan", "feb", "mar"]
MONTH_COLS_ACH = ["apr_ach", "may_ach", "jun_ach", "jul_ach", "aug_ach", "sep_ach", "oct_ach", "nov_ach", "dec_ach", "jan_ach", "feb_ach", "mar_ach"]
MONTH_LABELS = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]


@router.get("/acquisition-targets")
async def list_targets(
    financial_year: Optional[str] = None,
    type: Optional[str] = None,
    project_name: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = select(AcquisitionTarget).order_by(AcquisitionTarget.project_name)
    if financial_year:
        q = q.where(AcquisitionTarget.financial_year == financial_year)
    if type:
        q = q.where(AcquisitionTarget.type == type.upper())
    if project_name:
        q = q.where(AcquisitionTarget.project_name == project_name)
    r = await db.execute(q)
    return r.scalars().all()


@router.get("/acquisition-targets/export")
async def export_targets(
    project_name: Optional[str] = None,
    financial_year: Optional[str] = None,
    type: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = select(AcquisitionTarget).order_by(AcquisitionTarget.project_name, AcquisitionTarget.type, AcquisitionTarget.financial_year)
    if project_name:
        q = q.where(AcquisitionTarget.project_name == project_name)
    if financial_year:
        q = q.where(AcquisitionTarget.financial_year == financial_year)
    if type:
        q = q.where(AcquisitionTarget.type == type.upper())
    r = await db.execute(q)
    rows = r.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Acquisition Targets"

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="0B3D91", end_color="0B3D91", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Project", "Project Type", "BE/RE", "FY", "Basin",
               "Apr", "May", "Jun", "Jul", "Aug", "Sep",
               "Oct", "Nov", "Dec", "Jan", "Feb", "Mar",
               "Total", "Achieved", "%"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = thin_border

    for i, row in enumerate(rows, 2):
        vals = [
            row.project_name, row.project_type or "", row.type or "",
            row.financial_year or "", row.basin or "",
        ]
        total = 0
        for m in MONTH_COLS:
            v = getattr(row, m) or 0
            vals.append(v)
            total += v
        achieved = sum(getattr(row, m+"_ach") or 0 for m in MONTH_COLS)
        vals.append(total)
        vals.append(achieved)
        vals.append(round((achieved / total * 100) if total > 0 else 0, 1))

        for col, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=v)
            c.border = thin_border
            if col >= 6:
                c.alignment = Alignment(horizontal="right")

    for col in range(1, len(headers)+1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["E"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"acquisition_targets{'_'+project_name if project_name else ''}{'_'+financial_year if financial_year else ''}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/acquisition-targets/history/summary")
async def get_history_summary(
    financial_year: Optional[str] = None,
    month: Optional[str] = None,
    project_name: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = select(TargetMonthHistory).order_by(TargetMonthHistory.created_at.desc())
    q = q.join(AcquisitionTarget, TargetMonthHistory.target_id == AcquisitionTarget.id)
    if financial_year:
        q = q.where(AcquisitionTarget.financial_year == financial_year)
    if month:
        q = q.where(TargetMonthHistory.month == month)
    if project_name:
        q = q.where(AcquisitionTarget.project_name == project_name)
    r = await db.execute(q)
    return r.scalars().all()


@router.get("/acquisition-targets/financial-years")
async def list_financial_years(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    r = await db.execute(select(AcquisitionTarget.financial_year).distinct().order_by(AcquisitionTarget.financial_year))
    return [row[0] for row in r]


@router.get("/acquisition-targets/{target_id}")
async def get_target(target_id: int, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    t = await db.get(AcquisitionTarget, target_id)
    if not t:
        raise HTTPException(404, "Target not found")
    return t


@router.post("/acquisition-targets")
async def create_target(data: dict, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    t = AcquisitionTarget(**{k: v for k, v in data.items() if hasattr(AcquisitionTarget, k)})
    db.add(t)
    await db.commit()
    await db.refresh(t)
    return t


@router.put("/acquisition-targets/{target_id}")
async def update_target(target_id: int, data: dict, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    t = await db.get(AcquisitionTarget, target_id)
    if not t:
        raise HTTPException(404, "Target not found")
    role_name = user.role.name if user.role else "viewer"
    if t.approved:
        password = data.pop("_unlock_password", None)
        if role_name != "admin" or not password:
            raise HTTPException(403, "Target is approved. Only admin can edit after password verification.")
        if not verify_password(password, user.password_hash):
            raise HTTPException(400, "Incorrect admin password")
        data.pop("approved", None)
        data.pop("approved_by", None)
    MONTH_COLS = ["apr","may","jun","jul","aug","sep","oct","nov","dec","jan","feb","mar"]
    MONTH_COLS_ACH = [m+"_ach" for m in MONTH_COLS]
    histories = []
    for k, v in data.items():
        if hasattr(t, k):
            old = getattr(t, k)
            if k in MONTH_COLS and old != v:
                histories.append({"month": k, "field": "target", "old_value": old, "new_value": v})
            elif k in MONTH_COLS_ACH and old != v:
                histories.append({"month": k[:-4], "field": "achieved", "old_value": old, "new_value": v})
            setattr(t, k, v)
    for h in histories:
        db.add(TargetMonthHistory(
            target_id=target_id, month=h["month"], field=h["field"],
            old_value=h["old_value"], new_value=h["new_value"],
            changed_by=user.id,
        ))
    await db.commit()
    await db.refresh(t)
    return t


@router.post("/acquisition-targets/{target_id}/approve")
async def approve_target(target_id: int, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    role_name = user.role.name if user.role else "viewer"
    if role_name not in ("admin", "ops_manager"):
        raise HTTPException(403, "Only admin or ops_manager can approve")
    t = await db.get(AcquisitionTarget, target_id)
    if not t:
        raise HTTPException(404, "Target not found")
    t.approved = True
    t.approved_by = user.id
    await db.commit()
    return {"success": True, "approved": True}


@router.post("/acquisition-targets/{target_id}/unlock")
async def unlock_target(
    target_id: int,
    data: dict = Body(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    role_name = user.role.name if user.role else "viewer"
    if role_name != "admin":
        raise HTTPException(403, "Only admin can unlock")
    password = data.get("password", "")
    if not verify_password(password, user.password_hash):
        raise HTTPException(400, "Incorrect admin password")
    t = await db.get(AcquisitionTarget, target_id)
    if not t:
        raise HTTPException(404, "Target not found")
    t.approved = False
    t.approved_by = None
    await db.commit()
    return {"success": True, "approved": False}


@router.post("/acquisition-targets/{target_id}/request-approval")
async def request_approval(target_id: int, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    t = await db.get(AcquisitionTarget, target_id)
    if not t:
        raise HTTPException(404, "Target not found")
    if t.approved:
        raise HTTPException(400, "Target is already approved")
    if t.approval_requested:
        raise HTTPException(400, "Approval already requested for this target")
    role_name = user.role.name if user.role else "viewer"
    if role_name not in ("admin", "ops_manager", "data_creator"):
        raise HTTPException(403, "Only admin, ops_manager, or data_creator can request approval")
    t.approval_requested = True
    t.approval_requested_by = user.id
    await db.commit()
    return {"success": True, "approval_requested": True}


@router.post("/acquisition-targets/{target_id}/cancel-request")
async def cancel_approval_request(target_id: int, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    t = await db.get(AcquisitionTarget, target_id)
    if not t:
        raise HTTPException(404, "Target not found")
    if not t.approval_requested:
        raise HTTPException(400, "No approval request to cancel")
    if t.approved:
        raise HTTPException(400, "Target is already approved, cannot cancel request")
    role_name = user.role.name if user.role else "viewer"
    if role_name not in ("admin", "ops_manager") and t.approval_requested_by != user.id:
        raise HTTPException(403, "Only the requester or admin can cancel the request")
    t.approval_requested = False
    t.approval_requested_by = None
    await db.commit()
    return {"success": True, "approval_requested": False}


@router.get("/acquisition-targets/{target_id}/history")
async def get_target_history(
    target_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    r = await db.execute(
        select(TargetMonthHistory)
        .where(TargetMonthHistory.target_id == target_id)
        .order_by(TargetMonthHistory.created_at.desc())
    )
    return r.scalars().all()


@router.delete("/acquisition-targets/{target_id}", status_code=204)
async def delete_target(target_id: int, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    t = await db.get(AcquisitionTarget, target_id)
    if not t:
        raise HTTPException(404, "Target not found")
    await db.delete(t)
    await db.commit()


@router.get("/acquisition-targets/analytics/monthly")
async def monthly_analytics(
    financial_year: Optional[str] = None,
    project_name: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = select(AcquisitionTarget)
    if financial_year:
        q = q.where(AcquisitionTarget.financial_year == financial_year)
    if project_name:
        q = q.where(AcquisitionTarget.project_name == project_name)
    r = await db.execute(q)
    rows = r.scalars().all()

    result = {}
    for label in MONTH_LABELS:
        result[label] = {"be_target": 0, "be_achieved": 0, "re_target": 0, "re_achieved": 0}

    for row in rows:
        is_be = row.type == "BE"
        for i, month in enumerate(MONTH_COLS):
            label = MONTH_LABELS[i]
            tgt = getattr(row, month) or 0
            ach = getattr(row, MONTH_COLS_ACH[i]) or 0
            if is_be:
                result[label]["be_target"] += tgt
                result[label]["be_achieved"] += ach
            else:
                result[label]["re_target"] += tgt
                result[label]["re_achieved"] += ach

    return result


@router.get("/acquisition-targets/analytics/yearly")
async def yearly_analytics(
    financial_year: Optional[str] = None,
    project_name: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = select(AcquisitionTarget)
    if financial_year:
        q = q.where(AcquisitionTarget.financial_year == financial_year)
    if project_name:
        q = q.where(AcquisitionTarget.project_name == project_name)
    r = await db.execute(q)
    rows = r.scalars().all()

    be_target = 0
    be_achieved = 0
    re_target = 0
    re_achieved = 0

    for row in rows:
        if row.type == "BE":
            be_target += row.total or 0
            be_achieved += row.total_ach or 0
        else:
            re_target += row.total or 0
            re_achieved += row.total_ach or 0

    return {
        "be": {"target": be_target, "achieved": be_achieved},
        "re": {"target": re_target, "achieved": re_achieved},
    }


# ── Manpower Employees ──

@router.get("/manpower-employees")
async def list_manpower(
    section: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = select(ManpowerEmployee).order_by(ManpowerEmployee.section, ManpowerEmployee.sl_no)
    if section:
        q = q.where(ManpowerEmployee.section == section)
    r = await db.execute(q)
    emp_rows = r.scalars().all()
    emp_cpfs = {m.cpf_no for m in emp_rows if m.cpf_no}

    # Also fetch users from the system (exclude those already in manpower_employees)
    uq = select(User).options(selectinload(User.role)).where(User.is_active == True, User.cpf != None)
    ur = await db.execute(uq)
    system_users = ur.scalars().all()

    result = []
    for m in emp_rows:
        result.append({
            "id": m.id,
            "source": "seeded",
            "section": m.section,
            "basin": m.basin,
            "sl_no": m.sl_no,
            "cpf_no": m.cpf_no,
            "name": m.name,
            "designation": m.designation,
            "mobile": m.mobile,
            "level": m.level,
            "crc": m.crc,
            "assignment": m.assignment,
            "created_at": str(m.created_at) if m.created_at else None,
        })

    for su in system_users:
        if su.cpf and su.cpf not in emp_cpfs:
            result.append({
                "id": f"user_{su.id}",
                "source": "system",
                "section": su.section or "—",
                "basin": None,
                "sl_no": None,
                "cpf_no": su.cpf,
                "name": su.name,
                "designation": su.designation or "—",
                "mobile": None,
                "level": su.level or 0,
                "crc": None,
                "assignment": su.role.name if su.role else "",
                "created_at": str(su.created_at) if su.created_at else None,
            })

    return result


@router.get("/manpower-employees/sections")
async def list_manpower_sections(db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    r = await db.execute(select(ManpowerEmployee.section).distinct().order_by(ManpowerEmployee.section))
    return [row[0] for row in r]


@router.get("/manpower-employees/summary")
async def manpower_summary(db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    r = await db.execute(select(ManpowerEmployee))
    rows = r.scalars().all()
    by_section = {}
    by_level = {}
    for m in rows:
        sec = m.section or "Unknown"
        by_section[sec] = by_section.get(sec, 0) + 1
        lvl = m.level or "Unknown"
        by_level[lvl] = by_level.get(lvl, 0) + 1

    # Include system users (not already in manpower)
    ur = await db.execute(select(User).options(selectinload(User.role)).where(User.is_active == True, User.cpf != None))
    emp_cpfs = {m.cpf_no for m in rows if m.cpf_no}
    for su in ur.scalars().all():
        if su.cpf and su.cpf not in emp_cpfs:
            sec = su.section or "Unknown"
            by_section[sec] = by_section.get(sec, 0) + 1
            lvl = su.level or "Unknown"
            by_level[str(lvl)] = by_level.get(str(lvl), 0) + 1

    total = sum(by_section.values())
    return {"total": total, "by_section": by_section, "by_level": by_level}


@router.post("/manpower-employees")
async def create_manpower(data: dict, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    m = ManpowerEmployee(**{k: v for k, v in data.items() if hasattr(ManpowerEmployee, k)})
    db.add(m)
    await db.commit()
    await db.refresh(m)
    return m


@router.delete("/manpower-employees/{emp_id}", status_code=204)
async def delete_manpower(emp_id: int, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    m = await db.get(ManpowerEmployee, emp_id)
    if not m:
        raise HTTPException(404, "Manpower record not found")
    await db.delete(m)
    await db.commit()
